import openpyxl as xl
from openpyxl.chart import BarChart, Reference,BarChart3D
wb = xl.load_workbook('Groceriesptfinal.xlsx')                  # load xl workbook
sheet = wb['dATA']                                              # get the worksheet
cell = sheet['a1']                                              # get the cell row 1 and column 1
cell1 = sheet.cell(1, 1)                                         # get the cell
cell1 = sheet.cell(2, 6)
print(cell.value)                                               # get the value of cell
print(cell1.value)
print(sheet.max_row)                                            # get how many rows in the workbook
Eighth_cell = sheet.cell(2, 8)
Eighth_cell.value = cell1.value*.5
values = Reference(
            sheet,
            min_row=2,
            max_row=2,
            min_col=2,
            max_col=2

)
chart = BarChart3D()
chart.add_data(values)
sheet.add_chart(chart, 'j2')
wb.save('Groceriesptfinal.xlsx')
